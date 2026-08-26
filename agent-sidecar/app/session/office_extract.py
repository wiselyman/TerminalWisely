"""Extract plain text from local Office/PDF bytes (untrusted DATA)."""

from __future__ import annotations

import base64
import csv
import io
from typing import Any

# Soft caps — never load unbounded workbooks into the model context.
MAX_OFFICE_BYTES = 2 * 1024 * 1024
MAX_PDF_PAGES = 20
MAX_XLSX_ROWS = 200
MAX_XLSX_SHEETS = 3
MAX_TEXT_CHARS = 64 * 1024


def _decode_b64(data_base64: str) -> bytes | None:
    raw = (data_base64 or "").strip()
    if not raw:
        return None
    try:
        blob = base64.b64decode(raw, validate=False)
    except Exception:  # noqa: BLE001
        return None
    if len(blob) > MAX_OFFICE_BYTES:
        return None
    return blob


def extract_office_text(
    *,
    name: str,
    media_type: str | None,
    data_base64: str,
) -> dict[str, Any]:
    """Return {ok, text?, error?, truncated?}."""
    blob = _decode_b64(data_base64)
    if blob is None:
        return {
            "ok": False,
            "error": "office_bytes_missing_or_too_large",
            "name": name,
        }
    lower = (name or "").lower()
    mt = (media_type or "").lower()

    if lower.endswith(".pdf") or "pdf" in mt:
        return _extract_pdf(name, blob)
    if lower.endswith(".docx") or "wordprocessingml" in mt:
        return _extract_docx(name, blob)
    if lower.endswith(".xlsx") or "spreadsheetml" in mt:
        return _extract_xlsx(name, blob)
    if lower.endswith(".csv") or mt == "text/csv":
        return _extract_csv(name, blob)
    return {"ok": False, "error": "unsupported_office_type", "name": name}


def _clip(text: str) -> tuple[str, bool]:
    if len(text) <= MAX_TEXT_CHARS:
        return text, False
    keep = MAX_TEXT_CHARS // 2 - 32
    return text[:keep] + "\n...[truncated]...\n" + text[-keep:], True


def _extract_pdf(name: str, blob: bytes) -> dict[str, Any]:
    try:
        from pypdf import PdfReader
    except ImportError:
        return {"ok": False, "error": "pypdf_not_installed", "name": name}
    try:
        reader = PdfReader(io.BytesIO(blob))
        pages = reader.pages[:MAX_PDF_PAGES]
        parts: list[str] = []
        for i, page in enumerate(pages):
            parts.append(f"--- page {i + 1} ---\n{(page.extract_text() or '').strip()}")
        text, truncated = _clip("\n\n".join(parts).strip())
        if not text:
            return {"ok": False, "error": "pdf_no_extractable_text", "name": name}
        if len(reader.pages) > MAX_PDF_PAGES:
            truncated = True
        return {"ok": True, "text": text, "truncated": truncated, "name": name}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"pdf_extract_failed:{exc}", "name": name}


def _extract_docx(name: str, blob: bytes) -> dict[str, Any]:
    try:
        import docx  # python-docx
    except ImportError:
        return {"ok": False, "error": "python_docx_not_installed", "name": name}
    try:
        document = docx.Document(io.BytesIO(blob))
        paras = [p.text.strip() for p in document.paragraphs if p.text and p.text.strip()]
        text, truncated = _clip("\n".join(paras))
        if not text:
            return {"ok": False, "error": "docx_no_extractable_text", "name": name}
        return {"ok": True, "text": text, "truncated": truncated, "name": name}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"docx_extract_failed:{exc}", "name": name}


def _extract_xlsx(name: str, blob: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"ok": False, "error": "openpyxl_not_installed", "name": name}
    try:
        wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames[:MAX_XLSX_SHEETS]:
            ws = wb[sheet_name]
            parts.append(f"--- sheet: {sheet_name} ---")
            rows_out: list[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= MAX_XLSX_ROWS:
                    parts.append(f"...[truncated after {MAX_XLSX_ROWS} rows]...")
                    break
                cells = ["" if c is None else str(c) for c in row]
                if any(cells):
                    rows_out.append(",".join(cells))
            parts.extend(rows_out)
        text, truncated = _clip("\n".join(parts).strip())
        if not text:
            return {"ok": False, "error": "xlsx_no_extractable_text", "name": name}
        if len(wb.sheetnames) > MAX_XLSX_SHEETS:
            truncated = True
        return {"ok": True, "text": text, "truncated": truncated, "name": name}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"xlsx_extract_failed:{exc}", "name": name}


def _extract_csv(name: str, blob: bytes) -> dict[str, Any]:
    try:
        raw = blob.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(raw))
        rows: list[str] = []
        hit_cap = False
        for i, row in enumerate(reader):
            if i >= MAX_XLSX_ROWS:
                rows.append(f"...[truncated after {MAX_XLSX_ROWS} rows]...")
                hit_cap = True
                break
            rows.append(",".join(row))
        text, truncated = _clip("\n".join(rows))
        return {
            "ok": True,
            "text": text,
            "truncated": truncated or hit_cap,
            "name": name,
        }
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"csv_extract_failed:{exc}", "name": name}
